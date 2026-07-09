"""
Data Loader Module for Salesforce DCF Valuation Dashboard

This module reads financial data from the Excel model and structures it
for visualization in the Streamlit dashboard.
"""

import openpyxl
import pandas as pd
import numpy as np


class SalesforceDataLoader:
    """
    Loads and processes financial model data from Excel workbook.

    The Excel file contains:
    - Forecasting sheet: 3-statement model (Income Statement, Balance Sheet, Cash Flow)
    - Assumptions sheet: Scenario drivers and operational assumptions
    - DCF Valuation sheet: WACC, FCF projections, and intrinsic value calculation
    """

    def __init__(self, excel_path='5_years_data_with_valuation.xlsx'):
        """
        Initialize data loader and load Excel workbook.

        Args:
            excel_path (str): Path to the Excel model file
        """
        self.excel_path = excel_path
        # Load workbook with data_only=True to get calculated values instead of formulas
        self.wb = openpyxl.load_workbook(excel_path, data_only=True)

    def get_years(self):
        """Extract forecast years from the model (2027-2031)."""
        ws = self.wb['Forecasting']
        # Years are in row 2, columns 8-12 (2027-2031)
        years = [ws.cell(2, col).value for col in range(8, 13)]
        return years

    def get_revenue_data(self):
        """
        Extract revenue data for all years.

        Returns:
            dict: Years as keys, revenue values in millions
        """
        ws = self.wb['Forecasting']
        years = self.get_years()
        # Total Revenue is in row 4
        revenues = [ws.cell(4, col).value for col in range(8, 13)]
        return dict(zip(years, revenues))

    def get_margin_data(self):
        """
        Extract profitability margins over forecast period.

        Returns:
            DataFrame with gross margin, EBITDA margin, and net income margin
        """
        ws = self.wb['Forecasting']
        years = self.get_years()

        # Extract margin data (already calculated as percentages in the model)
        # Row 10: Gross margin
        # Row 18: EBITDA (need to calculate margin)
        # Row 28: Net income margin

        gross_margins = [ws.cell(10, col).value for col in range(8, 13)]
        revenues = [ws.cell(4, col).value for col in range(8, 13)]
        ebitdas = [ws.cell(18, col).value for col in range(8, 13)]
        net_margins = [ws.cell(28, col).value for col in range(8, 13)]

        # Calculate EBITDA margin
        ebitda_margins = [ebitda / revenue if revenue else 0
                         for ebitda, revenue in zip(ebitdas, revenues)]

        df = pd.DataFrame({
            'Year': years,
            'Gross Margin': gross_margins,
            'EBITDA Margin': ebitda_margins,
            'Net Income Margin': net_margins
        })

        return df

    def get_cash_data(self):
        """
        Extract cash and cash equivalents over forecast period.

        Returns:
            dict: Years as keys, cash balances in millions
        """
        ws = self.wb['Forecasting']
        years = self.get_years()
        # Row 33: CCE-marketable securities (cash position)
        cash_balances = [ws.cell(33, col).value for col in range(8, 13)]
        return dict(zip(years, cash_balances))

    def get_balance_sheet_checks(self):
        """
        Verify balance sheet integrity across forecast period.

        Returns:
            dict: Years as keys, check values (should be 0 for balanced)
        """
        ws = self.wb['Forecasting']
        years = self.get_years()
        # Row 74: Balance sheet check (Assets - Liabilities - Equity = 0)
        checks = [ws.cell(74, col).value for col in range(8, 13)]
        # Check if all values are effectively zero (within rounding tolerance)
        all_balanced = all(abs(c) < 0.01 if c is not None else True for c in checks)
        return {
            'years': years,
            'checks': checks,
            'all_balanced': all_balanced
        }

    def get_dcf_valuation(self):
        """
        Extract DCF valuation results.

        Returns:
            dict: Key valuation metrics including intrinsic value per share
        """
        ws = self.wb['DCF Valuation']

        # Extract key DCF inputs and outputs
        current_price = ws.cell(5, 3).value  # Row 5, Col C
        shares_outstanding = ws.cell(6, 3).value  # Row 6, Col C (in millions)
        wacc = ws.cell(23, 3).value  # Row 23, Col C
        enterprise_value = ws.cell(40, 2).value  # Row 40, Col B
        net_debt = ws.cell(41, 2).value  # Row 41, Col B
        equity_value = ws.cell(42, 2).value  # Row 42, Col B

        # Calculate intrinsic value per share
        intrinsic_value_per_share = equity_value / shares_outstanding

        # Calculate upside/downside vs. current market price
        upside_pct = ((intrinsic_value_per_share - current_price) / current_price) * 100

        return {
            'current_price': current_price,
            'shares_outstanding': shares_outstanding,
            'wacc': wacc,
            'enterprise_value': enterprise_value,
            'net_debt': net_debt,
            'equity_value': equity_value,
            'intrinsic_value_per_share': intrinsic_value_per_share,
            'upside_pct': upside_pct
        }

    def get_scenario_info(self):
        """
        Extract current scenario setting and available scenarios.

        Returns:
            dict: Current scenario and scenario descriptions
        """
        ws = self.wb['Assumptions']

        # Row 8, Col 2: Scenario selector (1=Base, 2=Bull, 3=Bear)
        scenario_selector = ws.cell(8, 2).value

        scenarios = {
            1: 'Base Case',
            2: 'Bull Case',
            3: 'Bear Case'
        }

        return {
            'current_scenario_id': scenario_selector,
            'current_scenario_name': scenarios.get(scenario_selector, 'Base Case'),
            'available_scenarios': scenarios
        }

    def calculate_revenue_cagr(self, start_year=2027, end_year=2031):
        """
        Calculate compound annual growth rate for revenue.

        Args:
            start_year (int): Starting year for CAGR calculation
            end_year (int): Ending year for CAGR calculation

        Returns:
            float: CAGR as a decimal (e.g., 0.088 for 8.8%)
        """
        revenue_data = self.get_revenue_data()
        start_revenue = revenue_data[start_year]
        end_revenue = revenue_data[end_year]
        n_years = end_year - start_year

        cagr = (end_revenue / start_revenue) ** (1/n_years) - 1
        return cagr

    def get_kpi_summary(self):
        """
        Generate summary KPIs for dashboard header tiles.

        Returns:
            dict: Key performance indicators for dashboard display
        """
        revenue_data = self.get_revenue_data()
        dcf_data = self.get_dcf_valuation()
        revenue_cagr = self.calculate_revenue_cagr()

        # Get FY2031 revenue and net income
        fy2031_revenue = revenue_data[2031]

        ws = self.wb['Forecasting']
        # Row 27: Net income, Column 12 for 2031
        fy2031_net_income = ws.cell(27, 12).value

        # Row 28: Net income margin for 2031
        fy2031_margin = ws.cell(28, 12).value

        return {
            'fy2031_revenue': fy2031_revenue,
            'fy2031_net_income': fy2031_net_income,
            'fy2031_margin': fy2031_margin,
            'revenue_cagr': revenue_cagr,
            'intrinsic_value': dcf_data['intrinsic_value_per_share'],
            'current_price': dcf_data['current_price'],
            'upside_pct': dcf_data['upside_pct'],
            'wacc': dcf_data['wacc']
        }

    def get_assets_liabilities_data(self):
        """
        Extract current assets and current liabilities over forecast period.

        Returns:
            DataFrame with current assets and current liabilities
        """
        ws = self.wb['Forecasting']
        years = self.get_years()

        # Row 39: Total current assets
        current_assets = [ws.cell(39, col).value for col in range(8, 13)]

        # Row 59: Total current liabilities
        current_liabilities = [ws.cell(59, col).value for col in range(8, 13)]

        df = pd.DataFrame({
            'Year': years,
            'Current Assets': current_assets,
            'Current Liabilities': current_liabilities
        })

        return df

    def get_assets_breakdown(self, year=2031):
        """
        Get detailed breakdown of current assets for a specific year.

        Args:
            year (int): Year to get breakdown for (default 2031)

        Returns:
            dict: Asset components with amounts and percentages
        """
        ws = self.wb['Forecasting']
        years = self.get_years()
        col_idx = 8 + years.index(year)  # Find column for the year

        # Current Asset components (rows 33-38)
        cce = ws.cell(33, col_idx).value or 0  # Cash & Cash Equivalents
        accounts_receivable = ws.cell(34, col_idx).value or 0
        inventory = ws.cell(35, col_idx).value or 0
        prepaid = ws.cell(36, col_idx).value or 0
        other_current = ws.cell(37, col_idx).value or 0

        total_current_assets = ws.cell(39, col_idx).value or 1

        breakdown = [
            {'name': 'Cash & Equivalents', 'amount': cce, 'pct': (cce / total_current_assets) * 100 if total_current_assets else 0},
            {'name': 'Accounts Receivable', 'amount': accounts_receivable, 'pct': (accounts_receivable / total_current_assets) * 100 if total_current_assets else 0},
            {'name': 'Inventory', 'amount': inventory, 'pct': (inventory / total_current_assets) * 100 if total_current_assets else 0},
            {'name': 'Prepaid Expenses', 'amount': prepaid, 'pct': (prepaid / total_current_assets) * 100 if total_current_assets else 0},
            {'name': 'Other Current Assets', 'amount': other_current, 'pct': (other_current / total_current_assets) * 100 if total_current_assets else 0},
        ]

        # Filter out zero values and sort by amount
        breakdown = [item for item in breakdown if item['amount'] > 0]
        breakdown.sort(key=lambda x: x['amount'], reverse=True)

        return {
            'total': total_current_assets,
            'components': breakdown
        }

    def get_liabilities_breakdown(self, year=2031):
        """
        Get detailed breakdown of current liabilities for a specific year.

        Args:
            year (int): Year to get breakdown for (default 2031)

        Returns:
            dict: Liability components with amounts and percentages
        """
        ws = self.wb['Forecasting']
        years = self.get_years()
        col_idx = 8 + years.index(year)  # Find column for the year

        # Current Liability components (rows 53-58)
        accounts_payable = ws.cell(53, col_idx).value or 0
        accrued_expenses = ws.cell(54, col_idx).value or 0
        deferred_revenue = ws.cell(55, col_idx).value or 0
        current_debt = ws.cell(56, col_idx).value or 0
        other_current_liab = ws.cell(57, col_idx).value or 0

        total_current_liabilities = ws.cell(59, col_idx).value or 1

        breakdown = [
            {'name': 'Accounts Payable', 'amount': accounts_payable, 'pct': (accounts_payable / total_current_liabilities) * 100 if total_current_liabilities else 0},
            {'name': 'Accrued Expenses', 'amount': accrued_expenses, 'pct': (accrued_expenses / total_current_liabilities) * 100 if total_current_liabilities else 0},
            {'name': 'Deferred Revenue', 'amount': deferred_revenue, 'pct': (deferred_revenue / total_current_liabilities) * 100 if total_current_liabilities else 0},
            {'name': 'Current Debt', 'amount': current_debt, 'pct': (current_debt / total_current_liabilities) * 100 if total_current_liabilities else 0},
            {'name': 'Other Current Liabilities', 'amount': other_current_liab, 'pct': (other_current_liab / total_current_liabilities) * 100 if total_current_liabilities else 0},
        ]

        # Filter out zero values and sort by amount
        breakdown = [item for item in breakdown if item['amount'] > 0]
        breakdown.sort(key=lambda x: x['amount'], reverse=True)

        return {
            'total': total_current_liabilities,
            'components': breakdown
        }

    def get_comps_valuation_estimate(self):
        """
        Generate estimated trading comps valuations using typical SaaS multiples.

        Note: Since trading comps are not in the Excel model, we create estimates
        based on the company's financial profile and typical SaaS valuation ranges.

        Returns:
            dict: Estimated per-share values from different valuation methods
        """
        revenue_data = self.get_revenue_data()
        dcf_data = self.get_dcf_valuation()

        # Use FY2027 (NTM) revenue for forward multiples
        ntm_revenue = revenue_data[2027]

        ws = self.wb['Forecasting']
        # Row 91: Operating Cash Flow for 2027 (Column 8)
        ocf_2027 = ws.cell(91, 8).value

        # Calculate levered FCF (OCF - Capex)
        # Row 100: Capex for 2027
        capex_2027 = ws.cell(100, 8).value
        fcf_2027 = ocf_2027 + capex_2027  # Capex is already negative

        # Typical SaaS multiples (conservative mid-range estimates)
        # High-growth SaaS companies typically trade at 5-15x NTM revenue
        # and 20-40x FCF depending on growth and profitability
        ev_revenue_multiple = 6.5  # Conservative multiple for mature SaaS
        ev_fcf_multiple = 25  # Conservative FCF multiple

        # Calculate enterprise value from multiples
        ev_from_revenue = ntm_revenue * ev_revenue_multiple
        ev_from_fcf = fcf_2027 * ev_fcf_multiple

        # Convert to equity value (EV - Net Debt)
        net_debt = dcf_data['net_debt']
        shares = dcf_data['shares_outstanding']

        equity_value_revenue = ev_from_revenue - net_debt
        equity_value_fcf = ev_from_fcf - net_debt

        # Calculate per-share values
        value_per_share_revenue = equity_value_revenue / shares
        value_per_share_fcf = equity_value_fcf / shares

        return {
            'ev_revenue_multiple': ev_revenue_multiple,
            'ev_fcf_multiple': ev_fcf_multiple,
            'value_per_share_revenue': value_per_share_revenue,
            'value_per_share_fcf': value_per_share_fcf,
            'note': 'Estimated using typical SaaS multiples (6.5x NTM Rev, 25x FCF)'
        }
